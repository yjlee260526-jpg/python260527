import requests
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# 네이버 검색 URL
url = "https://search.naver.com/search.naver"

# 검색 파라미터
params = {
    'where': 'nexearch',
    'sm': 'top_hty',
    'fbm': '0',
    'ie': 'utf8',
    'query': '반도체'
}

# User-Agent 설정 (네이버 접속으로 봇 차단 우회)
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
}

def save_to_excel(news_data):
    """
    크롤링한 뉴스 데이터를 Excel 파일로 저장하는 함수
    """
    try:
        # 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "네이버뉴스"
        
        # 헤더 스타일 설정
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 헤더 행 추가
        headers = ['번호', '제목', '링크', '설명', '언론사', '게시시간']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 데이터 행 추가
        for row_num, news in enumerate(news_data, 2):
            ws.cell(row=row_num, column=1).value = news['번호']
            ws.cell(row=row_num, column=2).value = news['제목']
            ws.cell(row=row_num, column=3).value = news['링크']
            ws.cell(row=row_num, column=4).value = news['설명']
            ws.cell(row=row_num, column=5).value = news['언론사']
            ws.cell(row=row_num, column=6).value = news['게시시간']
            
            # 데이터 셀 스타일 설정
            for col_num in range(1, 7):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # 열 너비 설정
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        # 행 높이 설정
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 30
        
        # 파일 저장
        filename = "naver_result.xlsx"
        wb.save(filename)
        
        print(f"\n✓ 엑셀 파일 저장 완료: {filename}")
        print(f"✓ 저장된 뉴스 개수: {len(news_data)}개")
        
    except Exception as e:
        print(f"Excel 파일 저장 중 오류: {e}")

try:
    # 페이지 요청
    response = requests.get(url, params=params, headers=headers)
    response.encoding = 'utf-8'
    
    # BeautifulSoup으로 파싱
    soup = BeautifulSoup(response.text, 'html.parser')
    
    # 뉴스 항목들 찾기 (최상위 뉴스 섹션)
    # 클래스 'sds-comps-vertical-layout sds-comps-full-layout dsRYUB0DWactfGrxvczL'로 각 뉴스 항목 구성
    news_items = soup.find_all('div', {'class': 'sds-comps-vertical-layout sds-comps-full-layout dsRYUB0DWactfGrxvczL'})
    
    # 데이터를 저장할 리스트
    news_data = []
    
    if news_items:
        print(f"\n[검색어: 반도체] 총 {len(news_items)}개의 뉴스를 찾았습니다.\n")
        print("=" * 80)
        
        # 각 뉴스 항목 처리
        for idx, item in enumerate(news_items, 1):
            try:
                # 뉴스 제목 추출
                # 제목 링크는 data-heatmap-target=".tit"인 a 태그
                title_tag = item.find('a', {'data-heatmap-target': '.tit'})
                if title_tag:
                    title = title_tag.find('span').get_text(strip=True) if title_tag.find('span') else title_tag.get_text(strip=True)
                    link = title_tag.get('href', '#')
                else:
                    title = "제목 없음"
                    link = "#"
                
                # 뉴스 설명 추출
                # 설명은 data-heatmap-target=".body"인 a 태그 내 span
                desc_tag = item.find('a', {'data-heatmap-target': '.body'})
                if desc_tag:
                    desc_span = desc_tag.find('span')
                    description = desc_span.get_text(strip=True) if desc_span else desc_tag.get_text(strip=True)
                else:
                    description = "설명 없음"
                
                # 언론사 정보 추출
                # 언론사는 프로필 섹션의 sds-comps-profile-info-title-text 클래스
                press_tag = item.find('span', {'class': 'sds-comps-profile-info-title-text'})
                if press_tag:
                    press_link = press_tag.find('a')
                    press = press_link.get_text(strip=True) if press_link else press_tag.get_text(strip=True)
                else:
                    press = "언론사 정보 없음"
                
                # 게시 시간 추출
                # 시간 정보는 "3시간 전", "9시간 전" 형태로 표시
                time_spans = item.find_all('span', {'class': 'sds-comps-text-type-body2'})
                publish_time = "시간 정보 없음"
                for span in time_spans:
                    text = span.get_text(strip=True)
                    if '시간 전' in text or '분 전' in text or '일 전' in text:
                        publish_time = text
                        break
                
                # 리스트에 데이터 추가
                news_data.append({
                    '번호': idx,
                    '제목': title,
                    '링크': link,
                    '설명': description,
                    '언론사': press,
                    '게시시간': publish_time
                })
                
                # 출력
                print(f"[뉴스 {idx}]")
                print(f"제목: {title}")
                print(f"링크: {link}")
                print(f"설명: {description}")
                print(f"언론사: {press}")
                print(f"게시시간: {publish_time}")
                print("-" * 80)
                
            except Exception as e:
                print(f"뉴스 항목 {idx} 파싱 중 오류: {e}")
                continue
        
        # Excel 파일로 저장
        if news_data:
            save_to_excel(news_data)
    else:
        print("뉴스 항목을 찾을 수 없습니다.")
        print("페이지 구조가 변경되었을 수 있습니다.")
        
except requests.exceptions.RequestException as e:
    print(f"요청 오류: {e}")
except Exception as e:
    print(f"오류 발생: {e}")
