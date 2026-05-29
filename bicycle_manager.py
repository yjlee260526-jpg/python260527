import sys
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QLabel, QLineEdit, QSpinBox, QPushButton, QTableWidget, 
    QTableWidgetItem, QMessageBox, QGroupBox
)
from PyQt6.QtCore import Qt


class BicycleManager(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db_path = 'bicycle.db'
        self.init_database()
        self.init_ui()
        self.load_data()

    def init_database(self):
        """데이터베이스 초기화 및 샘플 데이터 삽입"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # 기존 테이블 삭제 (재실행 시 초기화)
        cursor.execute('DROP TABLE IF EXISTS Bycle')
        
        # Bycle 테이블 생성
        cursor.execute('''
            CREATE TABLE Bycle (
                id INTEGER PRIMARY KEY,
                name TEXT NOT NULL,
                price INTEGER NOT NULL,
                qty INTEGER NOT NULL
            )
        ''')
        
        # 샘플 데이터 100개 삽입
        sample_data = []
        bicycle_names = [
            '로드 바이크', '마운틴 바이크', '하이브리드 바이크', '픽시 바이크', '키즈 바이크',
            '투어링 바이크', '그래블 바이크', 'BMX 바이크', '전기 자전거', '시티 바이크',
            '접이식 자전거', '스포츠 바이크', '캐주얼 바이크', '레이싱 바이크', '빅토리안 바이크'
        ]
        
        prices = [150000, 200000, 180000, 250000, 80000, 220000, 190000, 120000, 500000, 140000]
        
        for i in range(1, 101):
            name = f"{bicycle_names[i % len(bicycle_names)]} - {i}"
            price = prices[i % len(prices)] + (i * 1000)
            qty = (i % 50) + 1
            sample_data.append((name, price, qty))
        
        cursor.executemany('INSERT INTO Bycle (name, price, qty) VALUES (?, ?, ?)', sample_data)
        
        conn.commit()
        conn.close()

    def init_ui(self):
        """UI 초기화"""
        self.setWindowTitle('🚲 자전거 관리 시스템')
        self.setGeometry(100, 100, 1000, 700)
        
        # 메인 위젯
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        
        # 레이아웃
        main_layout = QVBoxLayout()
        
        # 입력 그룹박스
        input_group = QGroupBox('자전거 정보 입력')
        input_layout = QHBoxLayout()
        
        # ID
        input_layout.addWidget(QLabel('ID:'))
        self.id_input = QSpinBox()
        self.id_input.setRange(1, 10000)
        input_layout.addWidget(self.id_input)
        
        # 이름
        input_layout.addWidget(QLabel('이름:'))
        self.name_input = QLineEdit()
        input_layout.addWidget(self.name_input)
        
        # 가격
        input_layout.addWidget(QLabel('가격:'))
        self.price_input = QSpinBox()
        self.price_input.setRange(0, 1000000)
        input_layout.addWidget(self.price_input)
        
        # 수량
        input_layout.addWidget(QLabel('수량:'))
        self.qty_input = QSpinBox()
        self.qty_input.setRange(0, 10000)
        input_layout.addWidget(self.qty_input)
        
        input_group.setLayout(input_layout)
        main_layout.addWidget(input_group)
        
        # 버튼 그룹박스
        button_group = QGroupBox('작업')
        button_layout = QHBoxLayout()
        
        # 추가 버튼
        add_btn = QPushButton('추가')
        add_btn.clicked.connect(self.add_bicycle)
        button_layout.addWidget(add_btn)
        
        # 수정 버튼
        update_btn = QPushButton('수정')
        update_btn.clicked.connect(self.update_bicycle)
        button_layout.addWidget(update_btn)
        
        # 삭제 버튼
        delete_btn = QPushButton('삭제')
        delete_btn.clicked.connect(self.delete_bicycle)
        button_layout.addWidget(delete_btn)
        
        # 검색 버튼
        search_btn = QPushButton('검색')
        search_btn.clicked.connect(self.search_bicycle)
        button_layout.addWidget(search_btn)
        
        # 전체 표시 버튼
        all_btn = QPushButton('전체 표시')
        all_btn.clicked.connect(self.load_data)
        button_layout.addWidget(all_btn)
        
        # 초기화 버튼
        reset_btn = QPushButton('입력창 초기화')
        reset_btn.clicked.connect(self.clear_inputs)
        button_layout.addWidget(reset_btn)
        
        # 엑셀로 출력 버튼
        excel_btn = QPushButton('엑셀로 출력')
        excel_btn.clicked.connect(self.export_to_excel)
        button_layout.addWidget(excel_btn)
        
        button_group.setLayout(button_layout)
        main_layout.addWidget(button_group)
        
        # 테이블 위젯
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(['ID', '이름', '가격', '수량'])
        self.table.horizontalHeader().setStretchLastSection(False)
        self.table.setColumnWidth(0, 50)
        self.table.setColumnWidth(1, 300)
        self.table.setColumnWidth(2, 150)
        self.table.setColumnWidth(3, 100)
        
        # 테이블 행 선택 시 입력창에 데이터 표시
        self.table.itemSelectionChanged.connect(self.on_row_selected)
        
        main_layout.addWidget(self.table)
        
        main_widget.setLayout(main_layout)

    def load_data(self):
        """모든 데이터 로드"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM Bycle ORDER BY id')
        rows = cursor.fetchall()
        conn.close()
        
        self.display_data(rows)

    def display_data(self, rows):
        """테이블에 데이터 표시"""
        self.table.setRowCount(len(rows))
        
        for row_idx, (id_val, name, price, qty) in enumerate(rows):
            self.table.setItem(row_idx, 0, QTableWidgetItem(str(id_val)))
            self.table.setItem(row_idx, 1, QTableWidgetItem(name))
            self.table.setItem(row_idx, 2, QTableWidgetItem(f"{price:,}"))
            self.table.setItem(row_idx, 3, QTableWidgetItem(str(qty)))

    def on_row_selected(self):
        """테이블 행 선택 시 입력창에 데이터 표시"""
        selected_rows = self.table.selectedIndexes()
        if selected_rows:
            row = selected_rows[0].row()
            id_val = int(self.table.item(row, 0).text())
            name = self.table.item(row, 1).text()
            price = int(self.table.item(row, 2).text().replace(',', ''))
            qty = int(self.table.item(row, 3).text())
            
            self.id_input.setValue(id_val)
            self.name_input.setText(name)
            self.price_input.setValue(price)
            self.qty_input.setValue(qty)

    def add_bicycle(self):
        """자전거 추가"""
        name = self.name_input.text().strip()
        price = self.price_input.value()
        qty = self.qty_input.value()
        
        if not name:
            QMessageBox.warning(self, '입력 오류', '자전거 이름을 입력해주세요.')
            return
        
        if price <= 0:
            QMessageBox.warning(self, '입력 오류', '가격을 0 이상으로 입력해주세요.')
            return
        
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('INSERT INTO Bycle (name, price, qty) VALUES (?, ?, ?)', 
                          (name, price, qty))
            conn.commit()
            conn.close()
            
            QMessageBox.information(self, '성공', '자전거 정보가 추가되었습니다.')
            self.clear_inputs()
            self.load_data()
        except Exception as e:
            QMessageBox.critical(self, '오류', f'추가 실패: {str(e)}')

    def update_bicycle(self):
        """자전거 정보 수정"""
        id_val = self.id_input.value()
        name = self.name_input.text().strip()
        price = self.price_input.value()
        qty = self.qty_input.value()
        
        if id_val <= 0:
            QMessageBox.warning(self, '입력 오류', 'ID를 선택 또는 입력해주세요.')
            return
        
        if not name:
            QMessageBox.warning(self, '입력 오류', '자전거 이름을 입력해주세요.')
            return
        
        if price <= 0:
            QMessageBox.warning(self, '입력 오류', '가격을 0 이상으로 입력해주세요.')
            return
        
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('UPDATE Bycle SET name = ?, price = ?, qty = ? WHERE id = ?',
                          (name, price, qty, id_val))
            
            if cursor.rowcount == 0:
                QMessageBox.warning(self, '수정 실패', f'ID {id_val}인 자전거를 찾을 수 없습니다.')
            else:
                conn.commit()
                QMessageBox.information(self, '성공', '자전거 정보가 수정되었습니다.')
                self.clear_inputs()
                self.load_data()
            
            conn.close()
        except Exception as e:
            QMessageBox.critical(self, '오류', f'수정 실패: {str(e)}')

    def delete_bicycle(self):
        """자전거 삭제"""
        id_val = self.id_input.value()
        
        if id_val <= 0:
            QMessageBox.warning(self, '입력 오류', 'ID를 선택 또는 입력해주세요.')
            return
        
        reply = QMessageBox.question(self, '확인', 
                                     f'ID {id_val}인 자전거를 삭제하시겠습니까?',
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute('DELETE FROM Bycle WHERE id = ?', (id_val,))
                
                if cursor.rowcount == 0:
                    QMessageBox.warning(self, '삭제 실패', f'ID {id_val}인 자전거를 찾을 수 없습니다.')
                else:
                    conn.commit()
                    QMessageBox.information(self, '성공', '자전거 정보가 삭제되었습니다.')
                    self.clear_inputs()
                    self.load_data()
                
                conn.close()
            except Exception as e:
                QMessageBox.critical(self, '오류', f'삭제 실패: {str(e)}')

    def search_bicycle(self):
        """자전거 검색"""
        name = self.name_input.text().strip()
        
        if not name:
            QMessageBox.warning(self, '입력 오류', '검색할 자전거 이름을 입력해주세요.')
            return
        
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM Bycle WHERE name LIKE ? ORDER BY id', 
                          (f'%{name}%',))
            rows = cursor.fetchall()
            conn.close()
            
            if rows:
                self.display_data(rows)
                QMessageBox.information(self, '검색 결과', f'{len(rows)}개의 결과를 찾았습니다.')
            else:
                QMessageBox.information(self, '검색 결과', '검색 결과가 없습니다.')
                self.load_data()
        except Exception as e:
            QMessageBox.critical(self, '오류', f'검색 실패: {str(e)}')

    def clear_inputs(self):
        """입력 필드 초기화"""
        self.id_input.setValue(0)
        self.name_input.clear()
        self.price_input.setValue(0)
        self.qty_input.setValue(0)
        self.name_input.setFocus()

    def export_to_excel(self):
        """데이터를 Excel 파일로 내보내기"""
        try:
            # 데이터베이스에서 모든 데이터 조회
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM Bycle ORDER BY id')
            rows = cursor.fetchall()
            conn.close()
            
            if not rows:
                QMessageBox.warning(self, '내보내기 실패', '저장된 데이터가 없습니다.')
                return
            
            # 엑셀 파일 생성
            wb = Workbook()
            ws = wb.active
            ws.title = '자전거 정보'
            
            # 헤더 스타일 설정
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 헤더 작성
            headers = ['ID', '이름', '가격', '수량']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 데이터 작성
            for row_idx, (id_val, name, price, qty) in enumerate(rows, 2):
                ws.cell(row=row_idx, column=1, value=id_val)
                ws.cell(row=row_idx, column=2, value=name)
                ws.cell(row=row_idx, column=3, value=price)
                ws.cell(row=row_idx, column=4, value=qty)
                
                # 데이터 셀 중앙 정렬
                for col in range(1, 5):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 열 너비 자동 조정
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 10
            
            # 파일 저장
            filename = 'bicycle_data.xlsx'
            wb.save(filename)
            
            QMessageBox.information(self, '성공', f'데이터가 "{filename}"으로 저장되었습니다.')
        except Exception as e:
            QMessageBox.critical(self, '오류', f'Excel 내보내기 실패: {str(e)}')


def main():
    app = QApplication(sys.argv)
    window = BicycleManager()
    window.show()
    sys.exit(app.exec())


if __name__ == '__main__':
    main()
